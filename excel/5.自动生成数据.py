import openpyxl

# 1. 设置路径
input_path = r'/Users/mimihouse/Desktop/python/data/OH.xlsx'
output_path = r'/Users/mimihouse/Desktop/python/data/OH_calculated.xlsx'

# 2. 加载文件
print(f"正在读取: {input_path}")
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# 3. 写表头
ws['C2'] = "浓度"

# 4. 循环计算并设置格式
print("正在计算并调整格式...")
for row in ws.iter_rows(min_row=3, min_col=2, max_col=2):
    cell_b = row[0]  # 获取 B列 (峰面积) 的单元格
    b_value = cell_b.value

    # ---------------------------------------------
    # 🌟 修改点 1: 设置 B列 (峰面积) 显示为两位小数
    # ---------------------------------------------
    cell_b.number_format = '0.00'

    # 确保是数字才计算
    if b_value is not None and isinstance(b_value, (int, float)):
        # 计算公式
        concentration = ((b_value - 0.0667) / 0.0221) / 0.35

        # 获取 C列 (浓度) 对应的单元格
        cell_c = ws.cell(row=cell_b.row, column=3)

        # 写入计算结果
        cell_c.value = concentration

        # ---------------------------------------------
        # 🌟 修改点 2: 设置 C列 (浓度) 显示为两位小数
        # ---------------------------------------------
        cell_c.number_format = '0.00'

    else:
        # 如果是空值，C列也留空
        ws.cell(row=cell_b.row, column=3).value = None

# 5. 保存
wb.save(output_path)
print("✅ 完成！现在峰面积和浓度都保留了两位小数。")