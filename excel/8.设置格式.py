import openpyxl
from openpyxl.styles import PatternFill

# 1. 设置文件路径
file_path = r'/Users/mimihouse/Desktop/python/data/pollution degradation.xlsx'

# 2. 加载 Excel
print(f"正在读取: {file_path}")
try:
    wb = openpyxl.load_workbook(file_path)
    # 检查 comparison 表是否存在
    if 'comparison' not in wb.sheetnames:
        print("❌ 错误: 找不到 'comparison' Sheet，请先运行之前的计算代码。")
        exit()
    ws = wb['comparison']
except Exception as e:
    print(f"❌ 读取文件失败: {e}")
    exit()

# 3. 定义“标红”的样式
# start_color='FFCCCC': 浅红色 (Light Red)，文字看起来更清晰
# 如果想要深红，可以改成 'FF0000'
red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

# 4. 自动寻找“降解率”在第几列
rate_col_index = None
header_row = ws[1] # 获取第一行表头

for cell in header_row:
    # 这里的 cell.value 可能是 "降解率"
    if cell.value and "降解率" in str(cell.value):
        rate_col_index = cell.column # 获取列号 (1, 2, 3...)
        break

if rate_col_index is None:
    print("❌ 错误: 在表头中没找到'降解率'这一列，请检查表格。")
    exit()

print(f"定位成功: '降解率' 在第 {rate_col_index} 列")

# 5. 循环每一行进行判断 (从第2行开始)
count = 0
for row in ws.iter_rows(min_row=2):
    # 获取这一行里“降解率”那个单元格
    # row 是一个元组，索引从 0 开始，所以要减 1
    rate_cell = row[rate_col_index - 1]
    val = rate_cell.value

    # 判断是否大于 0.9
    # isinstance 确保它是数字，防止读到空值或文字报错
    if isinstance(val, (int, float)) and val > 0.9:
        count += 1
        # 如果符合条件，把这一行的所有单元格都标红
        for cell in row:
            cell.fill = red_fill

# 6. 保存
output_path = file_path # 直接覆盖保存，或者改成新名字
try:
    wb.save(output_path)
    print(f"✅ 完成！共有 {count} 个污染物的降解率 > 0.9，已标红。")
    print(f"文件已保存至: {output_path}")
except PermissionError:
    print("❌ 保存失败：请先关闭 Excel 文件！")